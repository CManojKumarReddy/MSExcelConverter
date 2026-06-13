"""
Tests for the Azure Document Intelligence image path (main.convert_image_azure)
and the route's graceful fallback to Tesseract.

No network/credentials needed: the Azure client is monkeypatched with a fake
layout result so we test our grid-mapping and Excel output, not Azure itself.
"""
import sys
import types
from pathlib import Path

import openpyxl
import pytest

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))
import main  # noqa: E402


# ── Fakes mimicking the Azure SDK's layout result shape ────────────────────────
class _FakeCell:
    def __init__(self, r, c, content):
        self.row_index = r
        self.column_index = c
        self.content = content


class _FakeTable:
    def __init__(self, rows):
        self.row_count = len(rows)
        self.column_count = max(len(r) for r in rows)
        self.cells = [
            _FakeCell(ri, ci, val)
            for ri, row in enumerate(rows)
            for ci, val in enumerate(row)
        ]


class _FakeResult:
    def __init__(self, tables, content=""):
        self.tables = tables
        self.content = content


class _FakePoller:
    def __init__(self, result):
        self._result = result

    def result(self):
        return self._result


def _install_fake_azure(monkeypatch, result):
    """Make convert_image_azure believe Azure is installed + configured, returning `result`."""
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "https://fake.cognitiveservices.azure.com/")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "fake-key")

    # Fake azure.core.credentials and azure.ai.documentintelligence modules.
    cred_mod = types.ModuleType("azure.core.credentials")
    cred_mod.AzureKeyCredential = lambda key: ("cred", key)

    class _FakeClient:
        def __init__(self, endpoint, credential):
            pass

        def begin_analyze_document(self, model_id, body=None, **kwargs):
            # body is an AnalyzeDocumentRequest in real use; we ignore it here.
            return _FakePoller(result)

    di_mod = types.ModuleType("azure.ai.documentintelligence")
    di_mod.DocumentIntelligenceClient = _FakeClient

    # azure.ai.documentintelligence.models.AnalyzeDocumentRequest(bytes_source=...)
    models_mod = types.ModuleType("azure.ai.documentintelligence.models")

    class _FakeReq:
        def __init__(self, bytes_source=None, url_source=None):
            self.bytes_source = bytes_source
            self.url_source = url_source

    models_mod.AnalyzeDocumentRequest = _FakeReq

    monkeypatch.setitem(sys.modules, "azure.core.credentials", cred_mod)
    monkeypatch.setitem(sys.modules, "azure.ai.documentintelligence", di_mod)
    monkeypatch.setitem(sys.modules, "azure.ai.documentintelligence.models", models_mod)


def _load(out_name):
    return openpyxl.load_workbook(main.OUTPUTS_DIR / out_name)


def test_azure_table_maps_to_sheet(monkeypatch):
    table = _FakeTable([
        ["Name", "Phone"],
        ["viswa", "9963393260"],
        ["ram", "7353582225"],
    ])
    _install_fake_azure(monkeypatch, _FakeResult([table]))

    out = main.convert_image_azure(b"fakebytes", "azure_t", mode="single")
    ws = _load(out).active
    rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
    assert ["Name", "Phone"] in rows
    assert ["viswa", "9963393260"] in rows
    assert ["ram", "7353582225"] in rows


def test_azure_separate_mode_one_sheet_per_table(monkeypatch):
    t1 = _FakeTable([["A", "B"], ["1", "2"]])
    t2 = _FakeTable([["C", "D"], ["3", "4"]])
    _install_fake_azure(monkeypatch, _FakeResult([t1, t2]))

    out = main.convert_image_azure(b"x", "azure_sep", mode="separate")
    wb = _load(out)
    assert len(wb.sheetnames) == 2


def test_azure_unavailable_when_not_configured(monkeypatch):
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "")
    with pytest.raises(main.CloudOCRUnavailable):
        main.convert_image_azure(b"x", "azure_none")


def test_azure_no_tables_falls_back_to_text(monkeypatch):
    _install_fake_azure(monkeypatch, _FakeResult([], content="Just some text\nsecond line"))
    out = main.convert_image_azure(b"x", "azure_txt")
    ws = _load(out).active
    text = " ".join(
        str(c) for r in ws.iter_rows(values_only=True) for c in r if c
    )
    assert "Just some text" in text


def test_azure_configured_flag(monkeypatch):
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "https://x/")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "k")
    assert main._azure_configured() is True
    monkeypatch.setattr(main, "AZURE_DI_KEY", "")
    assert main._azure_configured() is False


def test_cloud_ocr_status_endpoint(monkeypatch):
    from fastapi.testclient import TestClient
    client = TestClient(main.app)

    # Nothing configured → null engine
    monkeypatch.setattr(main, "GEMINI_API_KEY", "")
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "")
    assert client.get("/api/cloud-ocr-status").json() == {"engine": None}

    # Azure configured (and no Gemini) → azure
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "https://x/")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "k")
    assert client.get("/api/cloud-ocr-status").json() == {"engine": "azure"}

    # Gemini takes precedence when both are set
    monkeypatch.setattr(main, "GEMINI_API_KEY", "g")
    assert client.get("/api/cloud-ocr-status").json() == {"engine": "gemini"}
