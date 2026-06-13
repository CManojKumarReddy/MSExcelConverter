"""Pytest configuration for the backend test suite.

Makes `main` importable (tests live in backend/tests/, main.py in backend/) and
provides shared helpers for running image conversions and inspecting the result.
"""
import sys
import shutil
import os
from pathlib import Path

import pytest

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

import main  # noqa: E402  (import after sys.path tweak)

FIXTURES = Path(__file__).parent / "fixtures"


def _tesseract_available() -> bool:
    if shutil.which("tesseract"):
        return True
    for p in (
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
    ):
        if os.path.isfile(p):
            return True
    return False


# Skip the whole OCR suite cleanly when Tesseract isn't installed.
requires_tesseract = pytest.mark.skipif(
    not _tesseract_available(),
    reason="Tesseract OCR engine not installed (see CLAUDE.md / README for install).",
)


def _rapidocr_available() -> bool:
    try:
        import rapidocr_onnxruntime  # noqa: F401
        return True
    except Exception:
        return False


# Skip RapidOCR-specific assertions when the local neural OCR isn't installed.
requires_rapidocr = pytest.mark.skipif(
    not _rapidocr_available(),
    reason="rapidocr-onnxruntime not installed (pip install rapidocr-onnxruntime).",
)


# OCR is slow (~10s/image), so convert each fixture only once per test session
# and reuse the result across the tests that inspect it.
_CONVERSION_CACHE: dict = {}


@pytest.fixture
def convert_image_to_rows():
    """Run main.convert_image on a fixture and return (rows, chosen_strategy).

    rows: list[list[str]] from the produced .xlsx (None cells -> "").
    chosen_strategy: the strategy name convert_image logged as the winner
                     (e.g. 'header_band', 'left_aligned', 'word_based', 'grid'),
                     or None if not captured.

    Results are memoised per fixture so the (expensive) OCR runs once per image.
    The `caplog` arg is accepted for call-site compatibility but the strategy is
    captured via a logging handler on the first (uncached) run.
    """
    import logging
    import re
    import openpyxl

    def _run(fixture_name: str, caplog=None) -> tuple[list[list[str]], str | None]:
        if fixture_name in _CONVERSION_CACHE:
            return _CONVERSION_CACHE[fixture_name]

        path = FIXTURES / fixture_name
        assert path.exists(), f"missing fixture: {path}"
        content = path.read_bytes()

        # Capture the "chose '<strategy>'" log line directly off the main logger.
        records: list[logging.LogRecord] = []
        handler = logging.Handler()
        handler.emit = records.append          # type: ignore[assignment]
        main_logger = logging.getLogger("main")
        prev_level = main_logger.level
        main_logger.setLevel(logging.INFO)
        main_logger.addHandler(handler)
        try:
            out_name = main.convert_image(content, Path(fixture_name).stem, fixture_name)
        finally:
            main_logger.removeHandler(handler)
            main_logger.setLevel(prev_level)

        chosen = None
        for rec in records:
            m = re.search(r"chose '([a-z_]+)'", rec.getMessage())
            if m:
                chosen = m.group(1)

        wb = openpyxl.load_workbook(main.OUTPUTS_DIR / out_name)
        ws = wb.active
        rows = [
            [("" if c is None else str(c)) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        result = (rows, chosen)
        _CONVERSION_CACHE[fixture_name] = result
        return result

    return _run


def all_text(rows) -> str:
    """Flatten all cell text into one lowercased string for token checks."""
    return " ".join(c for r in rows for c in r if c).lower()


def max_nonempty_cols(rows) -> int:
    """Largest number of non-empty cells in any single row."""
    return max((sum(1 for c in r if c.strip()) for r in rows), default=0)


@pytest.fixture
def convert_pdf_to_sheets():
    """Run main.convert_pdf on a fixture; return {sheet_name: rows}.

    rows: list[list[str]] per sheet (None cells -> ""). Memoised per (fixture, mode).
    """
    import openpyxl

    cache: dict = {}

    def _run(fixture_name: str, mode: str = "single") -> dict:
        key = (fixture_name, mode)
        if key in cache:
            return cache[key]
        path = FIXTURES / fixture_name
        assert path.exists(), f"missing fixture: {path}"
        out_name = main.convert_pdf(path.read_bytes(), Path(fixture_name).stem, mode=mode)
        wb = openpyxl.load_workbook(main.OUTPUTS_DIR / out_name)
        result = {
            sn: [["" if c is None else str(c) for c in row]
                 for row in wb[sn].iter_rows(values_only=True)]
            for sn in wb.sheetnames
        }
        cache[key] = result
        return result

    return _run
