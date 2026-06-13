"""
Tests for the Google Gemini image path (main.convert_image_gemini) and the
cloud-OCR engine selection / fallback.

No network/key needed: the google.genai client is monkeypatched to return canned
JSON, so we test our JSON→sheet mapping and routing, not Gemini itself.
"""
import sys
import json
import types
from pathlib import Path

import openpyxl
import pytest

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))
import main  # noqa: E402


class _FakeResponse:
    def __init__(self, text):
        self.text = text


def _install_fake_gemini(monkeypatch, response_text=None, *, per_model=None):
    """Make convert_image_gemini believe google-genai is installed + keyed.

    Pass `response_text` (same for any model) or `per_model`: a dict mapping model
    name → an outcome, where an outcome is a response string, an Exception to raise,
    OR a list of those consumed FIFO across successive calls (last entry repeats).
    Also mocks time.sleep so retry waits don't slow the tests.
    """
    monkeypatch.setattr(main, "GEMINI_API_KEY", "fake-key")
    monkeypatch.setattr("time.sleep", lambda *_a, **_k: None)

    calls = {}  # model -> next index into its sequence

    def _resolve(model):
        outcome = per_model.get(model)
        if isinstance(outcome, list):
            i = min(calls.get(model, 0), len(outcome) - 1)
            calls[model] = calls.get(model, 0) + 1
            outcome = outcome[i]
        if isinstance(outcome, BaseException):
            raise outcome
        if outcome is None:
            raise RuntimeError(f"no fake outcome configured for model {model}")
        return _FakeResponse(outcome)

    class _FakeModels:
        def generate_content(self, model=None, contents=None, config=None):
            if per_model is not None:
                return _resolve(model)
            return _FakeResponse(response_text)

    class _FakeClient:
        def __init__(self, api_key=None):
            self.models = _FakeModels()

    genai_mod = types.ModuleType("google.genai")
    genai_mod.Client = _FakeClient

    types_mod = types.ModuleType("google.genai.types")
    types_mod.Part = type("Part", (), {"from_bytes": staticmethod(lambda data, mime_type: ("part", mime_type))})
    types_mod.GenerateContentConfig = lambda **kw: ("cfg", kw)

    google_pkg = sys.modules.get("google") or types.ModuleType("google")
    google_pkg.genai = genai_mod
    genai_mod.types = types_mod

    monkeypatch.setitem(sys.modules, "google", google_pkg)
    monkeypatch.setitem(sys.modules, "google.genai", genai_mod)
    monkeypatch.setitem(sys.modules, "google.genai.types", types_mod)


def _load(out_name):
    return openpyxl.load_workbook(main.OUTPUTS_DIR / out_name)


def test_gemini_table_maps_to_sheet(monkeypatch):
    payload = json.dumps({"tables": [{"rows": [
        ["Name", "Phone"],
        ["viswa", "9963393260"],
        ["ram", "7353582225"],
    ]}]})
    _install_fake_gemini(monkeypatch, payload)

    out = main.convert_image_gemini(b"imgbytes", "gem_t", mode="single")
    ws = _load(out).active
    rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
    assert ["Name", "Phone"] in rows
    assert ["viswa", "9963393260"] in rows
    assert ["ram", "7353582225"] in rows


def test_gemini_separate_mode_one_sheet_per_table(monkeypatch):
    payload = json.dumps({"tables": [
        {"rows": [["A", "B"], ["1", "2"]]},
        {"rows": [["C", "D"], ["3", "4"]]},
    ]})
    _install_fake_gemini(monkeypatch, payload)
    out = main.convert_image_gemini(b"x", "gem_sep", mode="separate")
    wb = _load(out)
    assert len(wb.sheetnames) == 2


def test_gemini_tolerates_code_fences(monkeypatch):
    payload = "```json\n" + json.dumps({"tables": [{"rows": [["x", "y"]]}]}) + "\n```"
    _install_fake_gemini(monkeypatch, payload)
    out = main.convert_image_gemini(b"x", "gem_fence")
    ws = _load(out).active
    text = " ".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if c)
    assert "x" in text and "y" in text


def test_gemini_unavailable_when_not_configured(monkeypatch):
    monkeypatch.setattr(main, "GEMINI_API_KEY", "")
    with pytest.raises(main.CloudOCRUnavailable):
        main.convert_image_gemini(b"x", "gem_none")


def test_gemini_bad_json_raises(monkeypatch):
    _install_fake_gemini(monkeypatch, "not json at all")
    with pytest.raises(main.CloudOCRUnavailable):
        main.convert_image_gemini(b"x", "gem_bad")


def test_gemini_switches_model_on_429(monkeypatch):
    """429 on the primary model → automatically retried on the fallback model."""
    monkeypatch.setattr(main, "GEMINI_MODEL", "gemini-2.5-flash")
    monkeypatch.setattr(main, "GEMINI_FALLBACK_MODEL", "gemini-2.0-flash")
    payload = json.dumps({"tables": [{"rows": [["ok", "fallback"]]}]})
    _install_fake_gemini(monkeypatch, per_model={
        "gemini-2.5-flash": RuntimeError("429 RESOURCE_EXHAUSTED: quota exceeded"),
        "gemini-2.0-flash": payload,
    })
    out = main.convert_image_gemini(b"x", "gem_switch")
    ws = openpyxl.load_workbook(main.OUTPUTS_DIR / out).active
    text = " ".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if c)
    assert "fallback" in text


def test_gemini_per_minute_429_waits_and_retries_same_model(monkeypatch):
    """A per-minute rate limit must NOT switch models — wait and retry the same one."""
    monkeypatch.setattr(main, "GEMINI_MODEL", "gemini-2.5-flash")
    monkeypatch.setattr(main, "GEMINI_FALLBACK_MODEL", "gemini-2.0-flash")
    ok = json.dumps({"tables": [{"rows": [["retried", "ok"]]}]})
    _install_fake_gemini(monkeypatch, per_model={
        # first call: per-minute 429 with a suggested delay; second call: success
        "gemini-2.5-flash": [
            RuntimeError("429 RESOURCE_EXHAUSTED ... "
                         "GenerateRequestsPerMinutePerProjectPerModel-FreeTier. "
                         "Please retry in 2s"),
            ok,
        ],
        # if it wrongly switched here, this model is limit:0 and would hard-fail
        "gemini-2.0-flash": RuntimeError("429 ... limit: 0 ... gemini-2.0-flash"),
    })
    out = main.convert_image_gemini(b"x", "gem_permin")
    ws = openpyxl.load_workbook(main.OUTPUTS_DIR / out).active
    text = " ".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if c)
    assert "retried" in text  # succeeded on 2.5-flash, never used the dead fallback


def test_gemini_per_day_429_switches_model(monkeypatch):
    """A per-DAY quota is hard for this model → switch to the fallback model."""
    monkeypatch.setattr(main, "GEMINI_MODEL", "gemini-2.5-flash")
    monkeypatch.setattr(main, "GEMINI_FALLBACK_MODEL", "gemini-2.0-flash")
    ok = json.dumps({"tables": [{"rows": [["from", "fallback"]]}]})
    _install_fake_gemini(monkeypatch, per_model={
        "gemini-2.5-flash": RuntimeError("429 RESOURCE_EXHAUSTED ... "
                                         "GenerateRequestsPerDayPerProjectPerModel-FreeTier"),
        "gemini-2.0-flash": ok,
    })
    out = main.convert_image_gemini(b"x", "gem_perday")
    ws = openpyxl.load_workbook(main.OUTPUTS_DIR / out).active
    text = " ".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if c)
    assert "fallback" in text


def test_gemini_all_models_429_raises(monkeypatch):
    monkeypatch.setattr(main, "GEMINI_MODEL", "gemini-2.5-flash")
    monkeypatch.setattr(main, "GEMINI_FALLBACK_MODEL", "gemini-2.0-flash")
    _install_fake_gemini(monkeypatch, per_model={
        "gemini-2.5-flash": RuntimeError("429 quota exceeded"),
        "gemini-2.0-flash": RuntimeError("429 quota exceeded"),
    })
    with pytest.raises(main.CloudOCRUnavailable):
        main.convert_image_gemini(b"x", "gem_all429")


def test_gemini_permanent_error_does_not_try_fallback(monkeypatch):
    """A non-quota, non-transient error (e.g. 400) aborts immediately."""
    monkeypatch.setattr(main, "GEMINI_MODEL", "gemini-2.5-flash")
    monkeypatch.setattr(main, "GEMINI_FALLBACK_MODEL", "gemini-2.0-flash")
    calls = {"n": 0}

    class _FakeModels:
        def generate_content(self, model=None, contents=None, config=None):
            calls["n"] += 1
            raise RuntimeError("400 INVALID_ARGUMENT: bad request")

    class _FakeClient:
        def __init__(self, api_key=None):
            self.models = _FakeModels()

    monkeypatch.setattr(main, "GEMINI_API_KEY", "fake-key")
    genai_mod = types.ModuleType("google.genai")
    genai_mod.Client = _FakeClient
    types_mod = types.ModuleType("google.genai.types")
    types_mod.Part = type("Part", (), {"from_bytes": staticmethod(lambda data, mime_type: ("p", mime_type))})
    types_mod.GenerateContentConfig = lambda **kw: ("cfg", kw)
    genai_mod.types = types_mod
    google_pkg = sys.modules.get("google") or types.ModuleType("google")
    google_pkg.genai = genai_mod
    monkeypatch.setitem(sys.modules, "google", google_pkg)
    monkeypatch.setitem(sys.modules, "google.genai", genai_mod)
    monkeypatch.setitem(sys.modules, "google.genai.types", types_mod)

    with pytest.raises(main.CloudOCRUnavailable):
        main.convert_image_gemini(b"x", "gem_400")
    assert calls["n"] == 1, "should not have tried the fallback model on a permanent error"


def test_success_message_is_generic_branded(monkeypatch):
    """The UI message is always the generic 'Converted by MSExcelConverter' —
    no engine/quota details leak to the user — and a file is still returned even
    when the cloud engine fails (Tesseract fallback)."""
    from fastapi.testclient import TestClient
    client = TestClient(main.app)

    monkeypatch.setattr(main, "GEMINI_API_KEY", "fake-key")  # engine == 'gemini'
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "")
    monkeypatch.setattr(
        main, "convert_image_gemini",
        lambda *a, **k: (_ for _ in ()).throw(
            main.CloudOCRUnavailable("429 RESOURCE_EXHAUSTED ... limit: 0")),
    )
    # Tesseract fallback returns a known file (don't actually OCR in this test).
    monkeypatch.setattr(main, "convert_image", lambda *a, **k: "fallback.xlsx")

    # 1x1 PNG
    png = bytes.fromhex(
        "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
        "890000000a49444154789c6300010000050001"
        "0d0a2db40000000049454e44ae426082"
    )
    resp = client.post(
        "/api/convert",
        data={"use_azure": "true"},
        files={"file": ("x.png", png, "image/png")},
    )
    body = resp.json()
    assert body["output_filename"] == "fallback.xlsx"
    assert body["message"] == "Converted by MSExcelConverter"


def test_cloud_engine_precedence(monkeypatch):
    monkeypatch.setattr(main, "GEMINI_API_KEY", "")
    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "")
    assert main._cloud_ocr_engine() is None

    monkeypatch.setattr(main, "AZURE_DI_ENDPOINT", "https://x/")
    monkeypatch.setattr(main, "AZURE_DI_KEY", "k")
    assert main._cloud_ocr_engine() == "azure"

    monkeypatch.setattr(main, "GEMINI_API_KEY", "g")
    assert main._cloud_ocr_engine() == "gemini"
